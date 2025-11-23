#!/usr/bin/env python3
"""Example usage of the YT-DeepReSearch Orchestrator."""

import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from orchestrator.workflow import Orchestrator
from orchestrator.job_queue import JobQueue


def example_1_create_and_process():
    """Example 1: Create a queue and process a single job."""
    print("=" * 60)
    print("Example 1: Create and Process a Single Job")
    print("=" * 60)
    
    # Create a job queue
    jq = JobQueue("example_queue.xlsx")
    
    # Add a topic
    topic = "The Science of Black Holes"
    jq.add_job(topic)
    print(f"✓ Added topic: {topic}\n")
    
    # Create orchestrator
    orchestrator = Orchestrator(
        job_queue_path="example_queue.xlsx",
        projects_dir="example_projects"
    )
    
    # Process the job
    print(f"Processing: {topic}\n")
    orchestrator.process_single_job(topic)
    
    # Check status
    status = jq.get_job_status(topic)
    print(f"\n✓ Job Status: {status['status']}")
    print(f"✓ Workspace: {status['output_path']}")


def example_2_batch_processing():
    """Example 2: Add multiple topics and process them."""
    print("\n" + "=" * 60)
    print("Example 2: Batch Processing")
    print("=" * 60)
    
    # Create a job queue with multiple topics
    jq = JobQueue("batch_queue.xlsx")
    
    topics = [
        "Evolution of the Internet",
        "The Human Immune System",
        "Renewable Energy Technologies"
    ]
    
    for topic in topics:
        jq.add_job(topic)
        print(f"✓ Added: {topic}")
    
    print(f"\nCreated queue with {len(topics)} topics")
    print("Run the following to process all jobs:")
    print("  cd src && python main.py --mode orchestrator --job-queue ../batch_queue.xlsx --projects-dir ../batch_projects")


def example_3_check_status():
    """Example 3: Check status of jobs in queue."""
    print("\n" + "=" * 60)
    print("Example 3: Check Job Status")
    print("=" * 60)
    
    import os
    from openpyxl import load_workbook
    
    queue_file = "example_queue.xlsx"
    
    if not os.path.exists(queue_file):
        print(f"Queue file not found: {queue_file}")
        return
    
    jq = JobQueue(queue_file)
    wb = load_workbook(queue_file)
    ws = wb.active
    
    print(f"\nStatus of jobs in {queue_file}:")
    print("-" * 60)
    
    for row_idx in range(2, ws.max_row + 1):
        topic = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=2).value
        
        if topic:
            status_icon = "✓" if status == "Completed" else "○"
            print(f"{status_icon} {topic:<40} [{status}]")
    
    wb.close()


def main():
    """Run examples."""
    print("\n" + "=" * 60)
    print("YT-DeepReSearch Orchestrator - Examples")
    print("=" * 60 + "\n")
    
    try:
        # Example 1: Single job
        example_1_create_and_process()
        
        # Example 2: Batch processing
        example_2_batch_processing()
        
        # Example 3: Check status
        example_3_check_status()
        
        print("\n" + "=" * 60)
        print("Examples completed!")
        print("=" * 60)
        print("\nCheck the following files:")
        print("  - example_queue.xlsx (job queue)")
        print("  - example_projects/ (workspaces)")
        print("  - batch_queue.xlsx (batch queue)")
        print("\nClean up with:")
        print("  rm -rf example_queue.xlsx example_projects/ batch_queue.xlsx")
        
    except KeyboardInterrupt:
        print("\n\nInterrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
