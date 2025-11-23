#!/usr/bin/env python3
"""Helper script to create and manage job queues."""

import sys
import argparse
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from orchestrator.job_queue import JobQueue


def create_queue(args):
    """Create a new job queue with topics."""
    jq = JobQueue(args.file)
    
    if args.topics:
        for topic in args.topics:
            if jq.add_job(topic):
                print(f"✓ Added: {topic}")
            else:
                print(f"✗ Already exists: {topic}")
    
    print(f"\nJob queue created: {args.file}")


def add_job(args):
    """Add a job to existing queue."""
    jq = JobQueue(args.file)
    
    if jq.add_job(args.topic):
        print(f"✓ Added: {args.topic}")
    else:
        print(f"✗ Topic already exists: {args.topic}")


def list_jobs(args):
    """List all jobs in the queue."""
    jq = JobQueue(args.file)
    
    from openpyxl import load_workbook
    
    wb = load_workbook(args.file)
    ws = wb.active
    
    print("\nJob Queue Status")
    print("=" * 100)
    print(f"{'Topic':<50} {'Status':<15} {'Duration (s)':<15}")
    print("=" * 100)
    
    for row_idx in range(2, ws.max_row + 1):
        topic = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=2).value
        duration = ws.cell(row=row_idx, column=5).value
        
        if topic:
            duration_str = f"{duration:.2f}" if duration else "-"
            print(f"{topic:<50} {status or 'N/A':<15} {duration_str:<15}")
    
    wb.close()
    print()


def reset_job(args):
    """Reset a job status to Pending."""
    jq = JobQueue(args.file)
    
    from openpyxl import load_workbook
    
    with jq.lock:
        wb = load_workbook(args.file)
        ws = wb.active
        
        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell_topic = ws.cell(row=row_idx, column=1).value
            
            if cell_topic == args.topic:
                # Reset to Pending
                ws.cell(row=row_idx, column=2, value="Pending")
                ws.cell(row=row_idx, column=3, value=None)
                ws.cell(row=row_idx, column=4, value=None)
                ws.cell(row=row_idx, column=5, value=None)
                ws.cell(row=row_idx, column=6, value=None)
                
                wb.save(args.file)
                print(f"✓ Reset {args.topic} to Pending")
                found = True
                break
        
        wb.close()
        
        if not found:
            print(f"✗ Topic not found: {args.topic}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Job Queue Management Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Create a new queue with topics
  %(prog)s create job_queue.xlsx --topics "Topic 1" "Topic 2" "Topic 3"
  
  # Add a job to existing queue
  %(prog)s add job_queue.xlsx "New Topic"
  
  # List all jobs
  %(prog)s list job_queue.xlsx
  
  # Reset a stuck job
  %(prog)s reset job_queue.xlsx "Topic Name"
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Command to execute')
    
    # Create command
    create_parser = subparsers.add_parser('create', help='Create a new job queue')
    create_parser.add_argument('file', help='Path to Excel file')
    create_parser.add_argument('--topics', nargs='+', help='Topics to add')
    
    # Add command
    add_parser = subparsers.add_parser('add', help='Add a job to the queue')
    add_parser.add_argument('file', help='Path to Excel file')
    add_parser.add_argument('topic', help='Topic to add')
    
    # List command
    list_parser = subparsers.add_parser('list', help='List all jobs')
    list_parser.add_argument('file', help='Path to Excel file')
    
    # Reset command
    reset_parser = subparsers.add_parser('reset', help='Reset a job to Pending')
    reset_parser.add_argument('file', help='Path to Excel file')
    reset_parser.add_argument('topic', help='Topic to reset')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    try:
        if args.command == 'create':
            create_queue(args)
        elif args.command == 'add':
            add_job(args)
        elif args.command == 'list':
            list_jobs(args)
        elif args.command == 'reset':
            reset_job(args)
    except Exception as e:
        print(f"\nError: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
