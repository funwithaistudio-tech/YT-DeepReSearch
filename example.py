#!/usr/bin/env python3
"""Example script demonstrating Phase 0 usage.

This script shows how to:
1. Add topics to the job queue
2. Run the orchestrator
3. Check topic status
"""

import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from orchestrator.job_queue import JobQueue
from config import JobStatus


def add_sample_topics():
    """Add sample topics to the job queue."""
    print("Adding sample topics to queue...")
    
    queue = JobQueue()
    
    topics = [
        ("The Evolution of Machine Learning", "Educational video topic"),
        ("Understanding Neural Networks", "Deep dive into neural networks"),
        ("The Future of Renewable Energy", "Climate and technology"),
    ]
    
    for topic, notes in topics:
        queue.add_topic(topic, notes)
        print(f"  ✓ Added: {topic}")
    
    print(f"\nTopics added to: {queue.queue_file}")
    print("\nRun the orchestrator with: cd src && python main.py")


def check_topic_status():
    """Check and display topic statuses."""
    print("Checking topic statuses...\n")
    
    queue = JobQueue()
    
    # Read all rows from Excel
    import openpyxl
    workbook = openpyxl.load_workbook(queue.queue_file)
    sheet = workbook.active
    
    print("=" * 80)
    print(f"{'Topic':<40} {'Status':<15} {'Duration':<12}")
    print("=" * 80)
    
    for row_idx in range(2, sheet.max_row + 1):
        topic = sheet.cell(row=row_idx, column=1).value
        status = sheet.cell(row=row_idx, column=2).value
        duration = sheet.cell(row=row_idx, column=5).value
        
        if topic:
            topic_short = topic[:37] + "..." if len(topic) > 40 else topic
            duration_str = f"{duration:.2f}s" if duration else "N/A"
            print(f"{topic_short:<40} {status:<15} {duration_str:<12}")
    
    print("=" * 80)
    workbook.close()


def main():
    """Main function."""
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python example.py add     - Add sample topics to queue")
        print("  python example.py status  - Check status of topics in queue")
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == "add":
        add_sample_topics()
    elif command == "status":
        check_topic_status()
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)


if __name__ == "__main__":
    main()
