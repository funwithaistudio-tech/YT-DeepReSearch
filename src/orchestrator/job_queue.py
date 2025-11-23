"""Excel-based job queue for managing research topics.

This module implements the JobQueue class for handling Excel operations as defined in
the Hybrid Hierarchical-GraphRAG System specification (Section 3).

The JobQueue manages topic status tracking using an Excel file with locking strategy
to prevent concurrent processing of the same topic.
"""

import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Dict, Any

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class JobQueue:
    """Manages Excel-based job queue for research topics.
    
    The JobQueue uses an Excel file to track topics and their processing status.
    It implements a locking strategy (Read Phase -> Write Phase with check) to
    prevent concurrent modifications.
    
    Attributes:
        excel_path: Path to the Excel file
        lock_timeout: Maximum time (seconds) a topic can be locked before considering it stale
    """
    
    # Column indices (1-based for openpyxl)
    COL_TOPIC = 1
    COL_STATUS = 2
    COL_TIMESTAMP_START = 3
    COL_TIMESTAMP_END = 4
    COL_DURATION_SECONDS = 5
    COL_QUALITY_SCORE = 6
    COL_ERROR_MESSAGE = 7
    COL_NOTES = 8
    
    # Status values
    STATUS_PENDING = "Pending"
    STATUS_IN_PROGRESS = "In_Progress"
    STATUS_COMPLETED = "Completed"
    STATUS_FAILED = "Failed"
    
    def __init__(self, excel_path: str = "job_queue/topics.xlsx", lock_timeout: int = 3600):
        """Initialize JobQueue.
        
        Args:
            excel_path: Path to the Excel file
            lock_timeout: Maximum lock time in seconds (default: 3600 = 1 hour)
        """
        self.excel_path = Path(excel_path)
        self.lock_timeout = lock_timeout
        self._ensure_excel_exists()
    
    def _ensure_excel_exists(self) -> None:
        """Ensure the Excel file exists with proper structure."""
        if not self.excel_path.exists():
            logger.info(f"Creating new Excel file: {self.excel_path}")
            self.excel_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "Topics"
            
            # Write headers
            headers = [
                "Topic",
                "Status",
                "Timestamp_Start",
                "Timestamp_End",
                "Duration_Seconds",
                "Quality_Score",
                "Error_Message",
                "Notes"
            ]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add some sample topics
            sample_topics = [
                ["The Science of Sleep", self.STATUS_PENDING, "", "", "", "", "", ""],
                ["Quantum Computing Basics", self.STATUS_PENDING, "", "", "", "", "", ""],
                ["Climate Change Solutions", self.STATUS_PENDING, "", "", "", "", "", ""],
            ]
            
            for row_idx, topic_data in enumerate(sample_topics, start=2):
                for col_idx, value in enumerate(topic_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(self.excel_path)
            logger.info(f"Created Excel file with sample topics")
    
    def get_next_pending_topic(self) -> Optional[str]:
        """Get the next topic with 'Pending' status.
        
        Reads the Excel file and finds the first row with "Pending" status.
        Also checks for stale "In_Progress" topics that exceed lock_timeout.
        
        Returns:
            Topic name if found, None if no pending topics
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            current_time = datetime.now(timezone.utc)
            
            # Iterate through rows (skip header)
            for row_idx in range(2, ws.max_row + 1):
                topic = ws.cell(row=row_idx, column=self.COL_TOPIC).value
                status = ws.cell(row=row_idx, column=self.COL_STATUS).value
                timestamp_start = ws.cell(row=row_idx, column=self.COL_TIMESTAMP_START).value
                
                if not topic:
                    continue
                
                # Check for pending topics
                if status == self.STATUS_PENDING:
                    logger.info(f"Found pending topic: {topic}")
                    return topic
                
                # Check for stale In_Progress topics
                if status == self.STATUS_IN_PROGRESS and timestamp_start:
                    try:
                        start_time = datetime.fromisoformat(timestamp_start)
                        elapsed = (current_time - start_time).total_seconds()
                        
                        if elapsed > self.lock_timeout:
                            logger.warning(
                                f"Found stale In_Progress topic: {topic} "
                                f"(locked for {elapsed:.0f}s, timeout: {self.lock_timeout}s)"
                            )
                            # Reset to Pending
                            ws.cell(row=row_idx, column=self.COL_STATUS, value=self.STATUS_PENDING)
                            ws.cell(row=row_idx, column=self.COL_TIMESTAMP_START, value="")
                            ws.cell(row=row_idx, column=self.COL_NOTES, value=f"Reset from stale lock at {current_time.isoformat()}")
                            wb.save(self.excel_path)
                            return topic
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Invalid timestamp for topic {topic}: {e}")
            
            logger.info("No pending topics found")
            return None
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return None
    
    def claim_topic(self, topic_name: str) -> bool:
        """Claim a topic for processing.
        
        Updates status to "In_Progress" and sets Timestamp_Start.
        Implements locking strategy: Read Phase -> Write Phase with check.
        
        Args:
            topic_name: Name of the topic to claim
            
        Returns:
            True if claim successful, False otherwise
        """
        try:
            # Read Phase: Load and find the topic
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            topic_row = None
            current_status = None
            
            for row_idx in range(2, ws.max_row + 1):
                topic = ws.cell(row=row_idx, column=self.COL_TOPIC).value
                if topic == topic_name:
                    topic_row = row_idx
                    current_status = ws.cell(row=row_idx, column=self.COL_STATUS).value
                    break
            
            if topic_row is None:
                logger.error(f"Topic not found: {topic_name}")
                return False
            
            # Check Phase: Verify status is still Pending
            if current_status != self.STATUS_PENDING:
                logger.warning(
                    f"Cannot claim topic '{topic_name}' - status is '{current_status}', not 'Pending'"
                )
                return False
            
            # Write Phase: Update status with check
            # NOTE: Race condition mitigation is limited with Excel files.
            # For production use, consider using a proper database with atomic operations
            # or implement file-level locking (e.g., fcntl on Unix, msvcrt on Windows).
            # The reload/recheck provides basic protection but is not fully atomic.
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Re-check status
            recheck_status = ws.cell(row=topic_row, column=self.COL_STATUS).value
            if recheck_status != self.STATUS_PENDING:
                logger.warning(
                    f"Cannot claim topic '{topic_name}' - status changed to '{recheck_status}' during claim"
                )
                return False
            
            # Update status
            timestamp_start = datetime.now(timezone.utc).isoformat()
            ws.cell(row=topic_row, column=self.COL_STATUS, value=self.STATUS_IN_PROGRESS)
            ws.cell(row=topic_row, column=self.COL_TIMESTAMP_START, value=timestamp_start)
            
            wb.save(self.excel_path)
            logger.info(f"Successfully claimed topic: {topic_name}")
            return True
            
        except Exception as e:
            logger.error(f"Error claiming topic '{topic_name}': {e}")
            return False
    
    def update_status(
        self,
        topic_name: str,
        status: str,
        timestamp_end: Optional[str] = None,
        duration_seconds: Optional[float] = None,
        quality_score: Optional[float] = None,
        error_message: Optional[str] = None,
        notes: Optional[str] = None,
    ) -> bool:
        """Update topic status and related fields.
        
        Args:
            topic_name: Name of the topic to update
            status: New status (Completed, Failed, etc.)
            timestamp_end: End timestamp (ISO format)
            duration_seconds: Duration in seconds
            quality_score: Quality score (0-100)
            error_message: Error message if failed
            notes: Additional notes
            
        Returns:
            True if update successful, False otherwise
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            topic_row = None
            for row_idx in range(2, ws.max_row + 1):
                topic = ws.cell(row=row_idx, column=self.COL_TOPIC).value
                if topic == topic_name:
                    topic_row = row_idx
                    break
            
            if topic_row is None:
                logger.error(f"Topic not found for update: {topic_name}")
                return False
            
            # Update fields
            ws.cell(row=topic_row, column=self.COL_STATUS, value=status)
            
            if timestamp_end:
                ws.cell(row=topic_row, column=self.COL_TIMESTAMP_END, value=timestamp_end)
            
            if duration_seconds is not None:
                ws.cell(row=topic_row, column=self.COL_DURATION_SECONDS, value=duration_seconds)
            
            if quality_score is not None:
                ws.cell(row=topic_row, column=self.COL_QUALITY_SCORE, value=quality_score)
            
            if error_message:
                ws.cell(row=topic_row, column=self.COL_ERROR_MESSAGE, value=error_message)
            
            if notes:
                ws.cell(row=topic_row, column=self.COL_NOTES, value=notes)
            
            wb.save(self.excel_path)
            logger.info(f"Updated status for topic '{topic_name}' to '{status}'")
            return True
            
        except Exception as e:
            logger.error(f"Error updating topic '{topic_name}': {e}")
            return False
    
    def add_topic(self, topic_name: str, notes: str = "") -> bool:
        """Add a new topic to the queue.
        
        Args:
            topic_name: Name of the topic to add
            notes: Optional notes
            
        Returns:
            True if addition successful, False otherwise
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Check if topic already exists
            for row_idx in range(2, ws.max_row + 1):
                existing_topic = ws.cell(row=row_idx, column=self.COL_TOPIC).value
                if existing_topic == topic_name:
                    logger.warning(f"Topic already exists: {topic_name}")
                    return False
            
            # Add new topic
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=self.COL_TOPIC, value=topic_name)
            ws.cell(row=new_row, column=self.COL_STATUS, value=self.STATUS_PENDING)
            ws.cell(row=new_row, column=self.COL_NOTES, value=notes)
            
            wb.save(self.excel_path)
            logger.info(f"Added new topic: {topic_name}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding topic '{topic_name}': {e}")
            return False
